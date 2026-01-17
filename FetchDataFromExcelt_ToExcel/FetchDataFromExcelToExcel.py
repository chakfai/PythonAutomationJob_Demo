from openpyxl import load_workbook, Workbook
import statistics

# Load source Excel
wb = load_workbook("source.xlsx")
ws = wb.active

# Define which columns to summarize
columns_to_summarize = {
    "B": "Age",
    "C": "Salary"
}

# Collect summaries
summary = {}

for col, label in columns_to_summarize.items():
    values = [cell.value for cell in ws[col] if isinstance(cell.value, (int, float))]
    if values:  # Only compute if we found numeric data
        summary[label] = {
            "count": len(values),
            "mean": statistics.mean(values),
            "min": min(values),
            "max": max(values)
        }
    else:
        summary[label] = {"error": "No numeric data found"}

# Write summary to new Excel
new_wb = Workbook()
new_ws = new_wb.active
new_ws.title = "Summary"

# Write header row
new_ws.append(["Column", "Metric", "Value"])

# Write each summary row
for label, stats in summary.items():
    for k, v in stats.items():
        new_ws.append([label, k, v])

new_wb.save("dest.xlsx")
print("已產生 dest.xlsx")